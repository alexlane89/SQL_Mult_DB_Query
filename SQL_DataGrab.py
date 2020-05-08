import pyodbc
from openpyxl import Workbook

#Establishing connection to SQL server
con = pyodbc.connect("DRIVER={SQL Server}; server=XX.XX.XX.XX;database=master;uid=user;pwd=pwd")

#command that executes queries on connected server
cur = con.cursor()

#gathering list of Databases on server
cur.execute("SELECT name FROM sys.databases")
db_name = [] #Initialize an array of database names
for row in cur:
    db_name.append(row.name)

#obtaining recipes run in each database on the server
recipes = []
for i in range(len(db_name)):
    arch = []
    sql = ("SELECT recipe FROM ["+ db_name[i]+ "].[dbo].[batchview]") #SQL query on each database in the db_names array
    cur.execute(sql)
    for row in cur:
        arch.append(row.recipe)
    recipes.append(arch)
    if db_name[i] == 'DVHisDB':
        break

#closing connection to SQL
cur.close()
con.close()

#Opening Excel workbook and pasting values
wb = Workbook()
ws = wb.active
for i in range(len(recipes)):
    #Giving Column Headers
    ws.cell(row=1,column = 1+i).value = db_name[i]
    for j in range(len(recipes[i])):
        #Pasting values and filtering for only meaningful data
        if len(str(recipes[i][j]).split('\\')) > 1: #there were a few data points with the value none so conditional is just to make the code work if there is nothing that needs filtering
            ws.cell(row=j+2,column = 1+i).value = str(recipes[i][j]).split('\\')[-1] #Split function just divides data into peices split by '\' and [-1] indexes for after last backslash
        else:
            ws.cell(row=j+2,column = 1+i).value = str(recipes[i][j])
wb.save('sample2.xlsx')
